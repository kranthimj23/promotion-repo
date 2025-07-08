import pandas as pd
from openpyxl import load_workbook
import os
import shutil
import subprocess
import sys
 
def load_differences(file_path):
    """Load differences from an Excel file."""
    return pd.read_excel(file_path)
 
def get_column_index(sheet, header):
    """Get the column index of a given header."""
    for cell in sheet[1]:  # Assuming headers are in the first row
        if cell.value == header:
            return cell.column  # Return column index (1-based)
    return None
 
def remove_empty_columns(df):
    """
    Remove columns where only the first row has a value and all other rows are empty.
    """
    cols_to_drop = []
    for col in df.columns:
        # Check if first row has a value (not NaN or None)
        first_row_val = df.at[0, col] if len(df) > 0 else None
        if pd.isna(first_row_val) or first_row_val is None:
            continue  # Skip columns with empty first row
 
        # Check if all other rows are empty
        if len(df) > 1:
            rest_rows = df[col].iloc[1:]
            # print(col)
            if first_row_val == "name":
 
                print(first_row_val)
                print(rest_rows.isna().all())
            if rest_rows.isna().all():
                cols_to_drop.append(col)
                # print("helllllllllo")
                # print(col)
        else:
            # Only one row, so no other rows to check, don't drop
            pass
    # print(cols_to_drop)
    if cols_to_drop:
        df.drop(columns=cols_to_drop, inplace=True)
    
    return df
 
 
def get_row_index(sheet,object_id):
    rows =len(sheet)
    all_columns = sheet.columns.tolist()
    # for i in all_columns:
    #     print(i)
 
    # print(all_columns[0])
    print(rows)
    for ind in range(rows):
        obj_id = sheet.get('object_id', pd.Series([None]*len(sheet))).iloc[ind] if ind < len(sheet) else None
        obj_id1 = sheet.get('object_id_1', pd.Series([None]*len(sheet))).iloc[ind] if ind < len(sheet) else None
 
        if(obj_id == object_id or obj_id1 == object_id ):
            return ind
    return
 
def update_target_file(differences_df, target_file):
    """Update the target DataFrame based on differences."""
    workbook = load_excel_sheets(target_file)
    for _, row in differences_df.iterrows():
        sheets_name = row['Sheet Name']
        field = row['Field']
        object_id = row['Object Id']
        dr_value = row['SIT Value']
        Change = row['Change']
        # print(sheet_name)
        # print(field)
        print(Change)
 
        if sheets_name not in workbook.keys():
            # Create an empty DataFrame with 'object_id' column and the current field to update
            # You can customize columns as needed; here we start with object_id and the field
            columns = ['object_id']
            if field != 'object_id':
                columns.append(field)
            df = pd.DataFrame(columns=columns)
            workbook[sheets_name] = df
        else:
            df = workbook[sheets_name]
            # df = pd.read_excel(target_file, sheet_name=sheets_name, header=0)
            # print(len(df))
        if(Change == "Added" and field == "object_id"):
            index = get_row_index(df, object_id)
            
            if index is None:
                new_row = pd.DataFrame({col: [None] for col in df.columns})
                new_row['object_id'] = dr_value  # Update 'object_id'
                # new_row[field] = dr_value  # Update the specified field
                
                df = pd.concat([df, new_row], ignore_index=True)
 
        else:
            index = get_row_index(df, object_id)
            # if(sheets_name == "pull_subscription"):
                # print(df[len(df)])
                # print(len(df))
            if index is not None:
                # print("index VALUE : "+ str(index))
                # print("sheet VALUE : "+ str(sheets_name))
                # print("Field VALUE : "+ str(field))
                # print("object_id VALUE : "+ str(object_id))
 
                # print("DR VALUE : "+ str(dr_value))
                if dr_value is None or dr_value == "":
                    df.loc[index, field] = None
                else:    
                    df.loc[index, field] = dr_value  # Update the DataFrame
            else:
                print("sheet VALUE : "+ str(sheets_name))
                print("Field VALUE : "+ str(field))
                print("object_id VALUE : "+ str(object_id))
 
                print("DR VALUE : "+ str(dr_value))
                print("index value is none")
                # break
        # print(df)
        workbook[sheets_name] = df  # Update the workbook dictionary
 
    # for sheet_name, df in workbook.items():
    for sheet_name in workbook.keys():
        df = workbook[sheet_name]
        workbook[sheet_name] = remove_empty_columns(df)
 
    # Save the updated workbook to the Excel file
    with pd.ExcelWriter(target_file) as writer:
        for sheet_name, df in workbook.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
 
    # for _, row in differences_df.iterrows():
    #     sheet_name = row['Sheet Name']
    #     field = row['Field']
    #     # field_row = row['Field Row']
    #     object_id = row['Object Id']
    #     dr_value = row['DR Value']
        
        
    #     if sheet_name in workbook.keys():
    #         df = workbook[sheet_name]
    #         index = get_row_index(df,object_id )
    #         print(index)
    #         print(sheet_name)
    #         print(field)
    #         print(dr_value)
    #         workbook[sheet_name][field].iloc[index] = dr_value
            # col_index = get_column_index(workbook[sheet_name], field)
            
            
            # if field  in workbook[sheet_name].columns:
            #     print(field)
            # if field_row > 1 or col_index > 1:
                # if sheet_name == "jenkins_vm":
                    # print("worked")
                    # print(str(col_index)+" "+str(field_row) +" " +str(dr_value))
                # workbook[sheet_name].cell(row = field_row,column = col_index).value = dr_value  # Update with DR Value
                # print(workbook[sheet_name].cell(row = field_row ,column = col_index))
            
    return workbook
 
def save_updated_file(workbook, output_file):
    """Save the modified workbook back to an Excel file."""
    workbook.save(filename = output_file)
 
def load_excel_sheets(file_path):
    """Load all sheets from an Excel file into a dictionary."""
    return pd.read_excel(file_path, sheet_name=None,header=0)
 
def clone_repo(repo_url, branch_name, target_folder):
    try:
        if os.path.exists(target_folder):
            shutil.rmtree(target_folder)
        os.makedirs(target_folder)  # Create the target folder
    except Exception as e:
        print(f"Error creating folder '{target_folder}': {e}")
        return
    # Clone the specified branch into the target folder
    try:
        subprocess.run(
            ["git", "clone", "--branch", branch_name, repo_url, target_folder],
            check=True
        )
        print(f"Successfully cloned '{branch_name}' branch into '{target_folder}'.")
    except subprocess.CalledProcessError as e:
        print(f"Error cloning the repository: {e}")
 
def main():
 
    repo_url = sys.argv[1]  # input("Enter repo url")
    # repo_name = input("")
    src_repo = "https://github.com/kranthimj23/zdt-manager-src.git"
    # promote_branch_x_1 = input("Enter the branch containing the stable release: ")
    promote_branch_x = sys.argv[2]  # input("Enter the branch containing the updated files for release: ")
    # lenv = sys.argv[3]  #  input("enter lower env name")
    henv = sys.argv[3]  #  input("enter higher env name")
    # target_folder_x_1 = os.path.join(os.getcwd(),"promo_x_1")
    target_folder_x = os.path.join(os.getcwd(),"promo_x")
    src_folder = os.path.join(os.getcwd(),"src")
    clone_repo(src_repo, "zdt-application", src_folder)
    clone_repo(repo_url, promote_branch_x, target_folder_x)
 
    differences_file =  f'{target_folder_x}/helm-charts/{henv}-values/infra-values/release_note/infra_difference.xlsx'   # Update with your actual path
    target_file = f'{target_folder_x}/helm-charts/{henv}-values/infra-values/dataset/infra_sheet.xlsx'
    target_autotfvars =  f'{target_folder_x}/helm-charts/{henv}-values/infra-values/terraform.tfvars'
    # subprocess.run()
   # Specify path to existing target file
    # output_file = '/Users/m26395/Downloads/updatehigher-UAT1_drx.xlsx'  # Specify output path for new file
 
    # Load differences from Excel file
    differences = load_excel_sheets(differences_file)
    differences_df = differences["differences"]
 
 
    updated_workbook = update_target_file(differences_df, target_file)
  
    print(f"Updated infra sheet saved to {target_file}")
 
    result = subprocess.run(
                        ["python3.11", f"{src_folder}/backend/infra/cd/scripts/generate_tfvars.py" , target_file ,target_autotfvars, f"{src_folder}/backend/infra/backend/infra/cd/tmpl" ],
                        check=True,
                        capture_output=True,
                        text=True
                    )
    print(result.stdout)
    print("auto.tfvars has been updated")
 
 
    try:
        subprocess.run(['git', 'add', "."], cwd =target_folder_x, check=True, capture_output=True, text=True)
        status_result = subprocess.run(['git', 'status'], cwd =target_folder_x, check=True, capture_output=True, text=True)
        print(status_result.stdout)
        print(status_result.stderr)
        # Pull latest changes with rebase to avoid non-fast-forward errors
        subprocess.run(['git', 'config', 'user.email', 'kranthimj23@gmail.com'], cwd =target_folder_x ,check=True, timeout=30)
        subprocess.run(['git', 'config', 'user.name', 'kranthimj23'], cwd =target_folder_x, check=True, timeout=30)
        subprocess.run(['git', 'commit', '-m',
                    f'Pushing the updates into the branch: {promote_branch_x} {henv}  '], cwd =target_folder_x, check=True, capture_output=True, text=True)
        # subprocess.run(['git', 'pull', '--rebase', 'origin', sys.argv[2]], cwd=target_folder_x, check=True, capture_output=True, text=True)
        # print("Pulled latest changes successfully.")
        # Push changes
        push_result = subprocess.run(['git', 'push', 'origin', promote_branch_x], cwd=target_folder_x, check=True, capture_output=True, text=True)
        print("Push successful:")
        print(push_result.stdout)
    except subprocess.CalledProcessError as e:
        print("Git command failed!")
        print("Return code:", e.returncode)
        print("Command:", e.cmd)
        print("Output:", e.output)
        print("Error:", e.stderr)
        sys.exit(1)
 
 
 
 
 
 
if __name__ == "__main__":
    main()
 