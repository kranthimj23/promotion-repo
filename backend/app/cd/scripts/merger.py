import subprocess
import sys
import tempfile
import shutil
import os
import openpyxl
from openpyxl import Workbook, load_workbook
import json
import yaml
import stat
from filecmp import cmp
import datetime
from openpyxl.utils import get_column_letter
 
 
########################################################################################################################
#### To clone the meta sheet####
########################################################################################################################
 
def clone_repo_master(github_url, branch_name, target_folder):
    try:
        if os.path.exists(target_folder):
            shutil.rmtree(target_folder)
        os.makedirs(target_folder)  # Create the target folder
    except Exception as e:
        #print(f"Error creating folder '{target_folder}': {e}")
        return
 
    # Clone the specified branch into the target folder
    try:
        subprocess.run(
            ["git", "clone", "--branch", branch_name, github_url, target_folder],
            check=True, stdout=subprocess.DEVNULL
        )
        ##print(f"Successfully cloned '{branch_name}' branch into '{target_folder}'.")
    except subprocess.CalledProcessError as e:
        print(f"Error cloning the repository: {e}")
        
 
########################################################################################################################
 
def find_column_index(sheet, env_name):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == env_name:
            return col
    raise ValueError(f"Environment '{env_name}' not found in header")
 
def find_last_updated_branch(sheet, col_idx):
    for row in range(sheet.max_row, 1, -1):
        val = sheet.cell(row=row, column=col_idx).value
        if val is not None and val != 'X':
            return val, row
    raise ValueError(f"No branch found in column index {col_idx}")
 
def create_new_branch(base_branch, new_version=None):
    if new_version:
        # Assuming branch format like PF-release/1.0.0, keep prefix and append new_version
        prefix = "release"
        new_branch = f"{prefix}/{new_version}"
    else:
        new_branch = f"{base_branch}_promotion_branch"
    ##print(f"Created new branch '{new_branch}' from base branch '{base_branch}'")
    return new_branch
 
def update_excel_with_new_branch(file_path, sheet, lower_env, new_branch):
    lower_col = find_column_index(sheet, lower_env)
    max_row = sheet.max_row + 1  # next empty row
 
    # Write new branch name in lower_env column
    sheet.cell(row=max_row, column=lower_col, value=new_branch)
 
    # Mark other env columns as 'X'
    for col in range(1, sheet.max_column + 1):
        if col != lower_col:
            sheet.cell(row=max_row, column=col, value='X')
 
    # Save workbook
    sheet.parent.save(file_path)
    #print(f"Excel updated: new branch '{new_branch}' written to '{lower_env}' column at row {max_row}.")
 
def fetch_branches(file_path, lower_env, update_lower_env, new_branch_created, higher_env, new_version=None):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
 
    lower_col = find_column_index(sheet, lower_env)
    higher_col = find_column_index(sheet, higher_env)
 
    lower_branch, _ = find_last_updated_branch(sheet, lower_col)
    ##print("this is a lower_branch from python script", lower_branch)
    higher_branch, _ = find_last_updated_branch(sheet, higher_col)
    ##print(higher_branch)
 
    if lower_env == 'dev' and lower_branch!= 'X':
        if lower_branch == higher_branch:
            if not new_version:
                raise ValueError("new_version must be provided when lower_env is 'dev'")
            new_branch = create_new_branch(lower_branch, new_version)
            update_excel_with_new_branch(file_path, sheet, lower_env, new_branch)
            #print("lower branch == higher branch in dev ---> " , lower_branch, new_branch)
            new_branch_created = True
            return lower_branch, new_branch, update_lower_env, new_branch_created
 
        if lower_branch != higher_branch and higher_branch != 'X':
            #print("lower branch != higher branch in dev ---> " , higher_branch, lower_branch)
            return higher_branch, lower_branch, update_lower_env, new_branch_created
    else:
        if lower_env != 'dev' and lower_branch == higher_branch:
            new_branch = create_new_branch(lower_branch, new_version)
            update_lower_env = True
            new_branch_created = True
            #print(f"New branch is created values to be populated in Lower-env-->{update_lower_env}")
            update_excel_with_new_branch(file_path, sheet, 'dev', new_branch)
            return lower_branch, new_branch, update_lower_env, new_branch_created
        else:
            #print(f"Branches differ. Promoting using existing branch '{lower_branch}' in '{lower_env}'.")
            return higher_branch, lower_branch, update_lower_env, new_branch_created
 
 
########################################################################################################################
 
####
    #### To create a new branch ####
####
def clean_non_dev_folders(temp_dir):
    """Remove all files in environment folders"""
    env_folders = []
    for folder in os.listdir(temp_dir):
        item_path = os.path.join(temp_dir, f"helm-charts")
        if folder == "helm-charts":
            for i in os.listdir(item_path):
                if i.endswith('values'):    #and not i.startswith('dev'):
                    env_folders.append(i)
                    env_path = os.path.join(item_path, i)
 
                    if os.path.exists(env_path):
 
                        # Remove all files while preserving directory structure
                        for root, dirs, files in os.walk(env_path):
                            for file in files:
                                file_path = os.path.join(root, file)
                                os.remove(file_path)
                                readme_path = os.path.join(root, "readme.md")
                                with open(readme_path, 'w') as f:
                                    f.write(f"The files of {env_path} are stored here")
 
def create_github_branch(github_url, base_branch, new_branch):
    """
    Creates a new GitHub branch from an existing branch and updates meta-sheet.xlsx
    - Clears files in all *values folders except dev-values
    - Preserves directory structure
    - Updates meta-sheet in master branch
    """
    temp_dir = tempfile.mkdtemp()
    try:
        # Verify base branch exists
        ls_remote = subprocess.run(
            ['git', 'ls-remote', '--heads', github_url, base_branch],
            capture_output=True,
            text=True,
            check=True
        )
 
        if not ls_remote.stdout:
            raise ValueError(f"Base branch '{base_branch}' not found in repository")
 
        # Clone repository for new branch creation
        new_branch_dir = os.path.join(temp_dir, "new_branch")
        subprocess.run(
            ['git', 'clone', '--single-branch', '-b', base_branch, github_url, new_branch_dir],
            check=True,timeout = 30, stdout=subprocess.DEVNULL
        )
        # Create and switch to new branch
        subprocess.run(['git', 'checkout', '-b', new_branch], cwd=new_branch_dir, check=True, timeout=30)
        # Clean  environment folders
        clean_non_dev_folders(new_branch_dir)
        #print(f"Created '{new_branch}' --------successfully")
        # Commit and push cleaned branch
        subprocess.run(['git', 'config', 'user.email', 'kranthimj23@gmail.com'], cwd=new_branch_dir, check=True, timeout=30,)
        subprocess.run(['git', 'config', 'user.name', 'kranthimj23'], cwd=new_branch_dir, check=True, timeout=30)
        subprocess.run(['git', 'add', '.'], cwd=new_branch_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        subprocess.run(
            ['git', 'commit', '-m', f'Initialize {new_branch}: Clean  environment folders'],
            cwd=new_branch_dir, check=True, timeout=30, stdout=subprocess.DEVNULL
        )
        result = subprocess.run(['git', 'push', 'origin', new_branch], cwd=new_branch_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        #print(f"Created and cleaned branch '{new_branch}' successfully")
        # Update meta-sheet in master branch
        master_dir = os.path.join(temp_dir, "master")
        subprocess.run(['git', 'clone', '-b', 'master', github_url, master_dir], check=True, timeout=30, stdout=subprocess.DEVNULL)
 
        # Load and update Excel file
        excel_path = os.path.join(master_dir, 'meta-sheet.xlsx')
        wb = load_workbook(excel_path)
        ws = wb.active
 
        # Find 'dev' column
        headers = [cell.value for cell in ws[1]]
        dev_col = headers.index('dev') + 1
 
        # Find first empty row in dev column
        next_row = 2
        while ws.cell(row=next_row, column=dev_col).value:
            next_row += 1
 
        # Add new branch name
        ws.cell(row=next_row, column=dev_col).value = new_branch
 
        # Fill empty cells in other columns with 'X'
        for col in range(1, ws.max_column + 1):
            if col == dev_col:
                continue  # Skip dev column
            cell = ws.cell(row=next_row, column=col)
            if not cell.value:  # Check if cell is empty
                cell.value = 'X'
 
        wb.save(excel_path)
 
        # Commit and push changes
        subprocess.run(['git', 'config', 'user.email', 'kranthimj23@gmail.com'], cwd=master_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        subprocess.run(['git', 'config', 'user.name', 'kranthimj23'], cwd=master_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        subprocess.run(['git', 'add', 'meta-sheet.xlsx'], cwd=master_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        subprocess.run(['git', 'commit', '-m', f'Add {new_branch} to meta-sheet'], cwd=master_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        subprocess.run(['git', 'push', 'origin', 'master'], cwd=master_dir, check=True, timeout=30, stdout=subprocess.DEVNULL)
        #print(f"Updated meta-sheet.xlsx with branch '{new_branch}'")
 
    except subprocess.CalledProcessError as e:
        #print(f"Git operation failed: {e.stderr}")
        return False
    except Exception as e:
        #print(f"Error: {str(e)}")
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
 
    return True
 
########################################################################################################################
 
def main():
    repo_name = "promotion-repo"
    lower_env = sys.argv[1]                  # Change as needed
    higher_env = sys.argv[2]
    github_url = sys.argv[3]                # Change as needed
    new_version = sys.argv[4]              # Provide new version for dev branch creation
    target_folder = os.path.join(os.getcwd(), 'execution')
    os.makedirs(target_folder, exist_ok=True)
    
    github_token = os.getenv("GIT_TOKEN")
    if github_token and "github.com" in github_url:
        # Inject token into repo URL (safe for HTTPS GitHub URLs)
        if github_url.startswith("https://"):
            github_url = github_url.replace("https://", f"https://{github_token}@")
        else:
            raise ValueError("Unsupported repo_url format. Must start with https://")

    
    
    meta_sheet_file_path = os.path.join(target_folder,f"meta-sheet.xlsx")
    update_lower_env = False
    reverse_promotion = False
    new_branch_created = False
    clone_repo_master(github_url, "master", target_folder)
    prev_branch, present_branch, update_lower_env, new_branch_created = fetch_branches(meta_sheet_file_path, lower_env, update_lower_env, new_branch_created,higher_env, new_version)
    #print(f"Current branch in '{lower_env}': {prev_branch}")
    if new_branch_created:
        create_github_branch(github_url,prev_branch,present_branch)
        #print("new branch will be created")
 
    #print(prev_branch, present_branch, update_lower_env, new_branch_created)
    # else:
    #     #print("No new branch created; promoting using existing branch.")
    envs = []
 
    if update_lower_env:
        envs.append('dev')
        # envs.append(lower_env)
        envs.append(higher_env)
        reverse_promotion = True
    else:
        envs.append(lower_env)
        envs.append(higher_env)
 
    def on_rm_error(func, path, exc_info):
     os.chmod(path, stat.S_IWRITE)
     func(path)
 
    shutil.rmtree('execution', onerror=on_rm_error)
    x1 = prev_branch
    x2 = present_branch
    low = envs[0]
    high = envs[1]
    isNew = new_branch_created
    print(f"{x1}, {x2}, {low}, {high}, {isNew}")
 
if __name__ == "__main__":
    main()
########################################################################################################################
 
