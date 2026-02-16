import os
import re
import yaml
import json
import openpyxl
from openpyxl import load_workbook
import subprocess
import tempfile
import shutil
import sys

from git_helpers import (
    inject_git_token,
    is_base_branch_exists,
    clone_single_branch_and_checkout,
    clone_repo_and_checkout,
    configure_git_user,
    stage_commit_and_push,
    stage_specific_files_commit_and_push,
)
 
update_template= []
deleted_services = []
 
 
def get_sheets_with_values(excel_file, env):
    wb = load_workbook(excel_file, data_only=True)
    matching_sheets = ''
 
    for sheet_name in wb.sheetnames:
        print("Checking sheet:", sheet_name)
 
        if sheet_name in ("dev2", "Summary", "AQL", "SQL"):
            print("Skipping the sheet: ", sheet_name)
 
        else:
            ws = wb[sheet_name]
 
            # Find the column index of the "env-current value" header
            value_col_idx = None
            for cell in ws[1]:
                if cell.value == f"{env}-current value":
                    value_col_idx = cell.column  # 1-based index
                    print(f"Found column for '{env}-current value' in {sheet_name}: Column {value_col_idx}")
                    break
 
            if value_col_idx is not None:
                for row in ws.iter_rows(min_row=2, min_col=value_col_idx, max_col=value_col_idx):
                    cell_value = row[0].value
                    if cell_value is not None and str(cell_value).strip() != "":
                        matching_sheets = sheet_name
                        break  # No need to check further cells in this sheet
 
    if not matching_sheets:
        print("No values updated in any sheet for promotion.")
    else:
        print("Sheets with values:", matching_sheets)
 
    return matching_sheets
 
 
 
def read_yaml_files_to_json(folder_path):
    json_data = {}
 
    for filename in os.listdir(folder_path):
        if filename.endswith('.yaml') or filename.endswith('.yml'):
            yaml_file_path = os.path.join(folder_path, filename)
            with open(yaml_file_path, 'r') as f:
                yaml_content = yaml.safe_load(f)
                root_object = os.path.splitext(filename)[0]
                json_data[root_object] = yaml_content
    return json_data
 
def try_parse_json(value):
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return value


def handle_data_env(json_data, service_name, change_request, parsed_value):
    obj = json_data.setdefault(service_name, [])

    def get_name(val):
        if isinstance(val, dict):
            return val.get('name')
        return str(val)

    if change_request == 'add':
        if not any(isinstance(entry, dict) and entry.get('name') == get_name(parsed_value) for entry in obj):
            obj.append(parsed_value)
            print(f"Added to '{service_name}': {parsed_value}")
        else:
            print(f"Warning: Entry with name '{get_name(parsed_value)}' already exists in '{service_name}'.")

    elif change_request == 'modify':
        for entry in obj:
            if isinstance(entry, dict) and entry.get('name') == get_name(parsed_value):
                entry.update(parsed_value if isinstance(parsed_value, dict) else {})
                print(f"Modified entry in '{service_name}': {entry}")
                break
        else:
            print(f"Warning: Attempted to modify non-existent entry '{get_name(parsed_value)}' in '{service_name}'.")

    elif change_request == 'delete':
        json_data[service_name] = [entry for entry in obj if not (isinstance(entry, dict) and entry.get('name') == get_name(parsed_value))]
        print(f"Deleted entry from '{service_name}' with name '{get_name(parsed_value)}'.")

 
# def handle_data_env(json_data, service_name, change_request, parsed_value):
#     """Handles modifications for 'data' and 'env' services."""
#     obj = json_data.setdefault(service_name, [])
 
#     if change_request == 'add':
#         if not any(entry['name'] == parsed_value['name'] for entry in obj):
#             obj.append(parsed_value)
#             print(f"Added to '{service_name}': {parsed_value}")
#         else:
#             print(f"Warning: Entry with name '{parsed_value['name']}' already exists in '{service_name}'.")
 
#     elif change_request == 'modify':
#         for entry in obj:
#             if entry['name'] == parsed_value['name']:
#                 entry.update(parsed_value)
#                 print(f"Modified entry in '{service_name}': {entry}")
#                 break
#         else:
#             print(f"Warning: Attempted to modify non-existent entry '{parsed_value['name']}' in '{service_name}'.")
 
#     elif change_request == 'delete':
#         json_data[service_name] = [entry for entry in obj if entry['name'] != parsed_value['name']]
#         print(f"Deleted entry from '{service_name}' with name '{parsed_value['name']}'.")
 
# def insert_hardcoded_value(folder,service_list):
#     """
#     Inserts a hardcoded multi-line value in the 'env' section at the last-but-one position
#     if multiple values exist, or at the end if only one value is present.
#     """
#     temp_path = os.path.join(folder, "helm-charts", "templates", "deployment.yaml")
 
#     # Read the YAML file as text
#     with open(temp_path, 'r') as file:
#         text = file.read()
#     # Regex to locate the 'env:' section and capture its contents
#     env_pattern = r"(env:\s*\n(?:\s*\{\{.*\}\}\n?)+)"
#     # Find the env block
#     match = re.search(env_pattern, text)
#     if not match:
#         print("No 'env' section found.")
#         return text
 
#     # Extract the env block and split it into lines
#     env_block = match.group(1)
#     env_lines = env_block.strip().splitlines()
 
#     if env_lines[-1]:
#         indent = len(env_lines[-1]) - len((env_lines[-1]).strip())
#         indentation = " " * indent
#         print(indentation + ":indent")
 
#     if "{{- end }}" in env_lines[-1] and "{{- end }}" in env_lines[-2]:
#             env_lines[-1] = ""
#             print("end removed")
 
#     for service_file in service_list:
#         service_file = service_file.replace('-','_')
#         # Define the hardcoded block using the `service_file` variable
#         hardcoded_block = f"""{{{{- with .Values.env.{service_file} }}}}
#               {{{{- toYaml . | nindent 12 }}}}
#             {{{{- end }}}}"""
 
#         # Insert the hardcoded block at the last-but-one position, if possible
#         env_lines.append(indentation + hardcoded_block.strip())
 
#     env_lines.append(indentation+"{{- end}}")
#     env_lines.append(indentation+"\n")
 
#     # Reassemble the modified env block and replace it in the original text
#     modified_env_block = "\n".join(env_lines)
#     modified_text = text.replace(env_block, modified_env_block)
 
 
#     with open(temp_path, 'w') as file:
#         file.write(modified_text)
 
#     return modified_text


def insert_hardcoded_value(folder, service_list):
    """
    Inserts a hardcoded multi-line value in the 'env' section safely.
    Detects proper indentation and avoids duplicate {{ end }}.
    """
    temp_path = os.path.join(folder, "helm-charts", "templates", "deployment.yaml")

    with open(temp_path, "r") as f:
        lines = f.readlines()

    new_lines = []
    inside_env = False
    env_start_idx = None
    env_indent = ""
    block_depth = 0

    for i, line in enumerate(lines):
        stripped = line.strip()

        if not inside_env and stripped.startswith("env:"):
            inside_env = True
            env_start_idx = i
            env_indent = re.match(r"^(\s*)", line).group(1)
            new_lines.append(line)
            continue

        if inside_env:
            if re.search(r"{{[-]?\s*(with|if|range)\b", stripped):
                block_depth += 1
            elif re.search(r"{{[-]?\s*end\s*}}", stripped):
                block_depth -= 1

            # Detect end of env block (blank line or new YAML key at same indent)
            if block_depth <= 0 and stripped and not stripped.startswith("{{"):
                inside_env = False
                # Insert our hardcoded blocks here before moving on
                for service_file in service_list:
                    sf = service_file.replace("-", "_")
                    hardcoded_block = [
                        f"{env_indent}  {{- with .Values.env.{sf} }}\n",
                        f"{env_indent}    {{- toYaml . | nindent 12 }}\n",
                        f"{env_indent}  {{- end }}\n"
                    ]
                    new_lines.extend(hardcoded_block)
            new_lines.append(line)
        else:
            new_lines.append(line)

    with open(temp_path, "w") as f:
        f.writelines(new_lines)
 
 
def apply_changes_to_json(json_data, excel_file_path, sheet_name, lower_env, higher_env):
    print("Apply changes to json")
    wb = load_workbook(excel_file_path, data_only=True)
    ws = wb[sheet_name]
 
    # Read header row to map column names to indices
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[cell.value] = idx
 
    # Helper to get cell value by column name and row number
    def get_cell_value(row_num, col_name):
        col_idx = headers.get(col_name)
        if col_idx is None:
            return None
        return ws.cell(row=row_num, column=col_idx).value
 
    # Iterate over data rows starting from row 2
    for row_num in range(2, ws.max_row + 1):
        service_name = get_cell_value(row_num, 'Service name')
        change_request = get_cell_value(row_num, 'Change Request')
        key = get_cell_value(row_num, 'Key')
        comments = get_cell_value(row_num, 'Comment')
        le_cur = get_cell_value(row_num, f'{lower_env}-current value')
        le_prev = get_cell_value(row_num, f'{lower_env}-previous value')
        he_cur = get_cell_value(row_num, f'{higher_env}-current value')
        he_prev = get_cell_value(row_num, f'{higher_env}-previous value')
 
 
        # Handle addition and deletion of a root object
        if key is None or (isinstance(key, str) and key.strip() == ""):
            print(service_name, comments, "root user added/deleted")
 
            if comments.strip().lower() == "root object deleted":
                print("root object to be deleted found")
                if service_name and service_name in json_data:
                    print("True")
                    deleted_service = service_name
                    deleted_services.append(deleted_service)
                    del json_data[service_name]
                    print(f"Deleted root object: {service_name}")
                else:
                    print("Did not enter the loop", service_name, json_data[service_name])
                continue
 
 
            elif comments == "root object added":
                new_root_object = service_name
                if new_root_object not in json_data:
                    json_data[new_root_object] = {}
                    print(f"Added new root object: {new_root_object}")
                if he_cur is not None and (not (isinstance(he_cur, str) and he_cur.strip() == "")):
                    # Check if the value is a HYPERLINK and extract the path from it
                    if isinstance(he_cur, str) and he_cur.startswith('file'):
                        hyperlink_path = extract_hyperlink_path(he_cur)
                        hyperlink_path = rf'/{hyperlink_path}'
                        with open(hyperlink_path, 'r') as txt_file:
                            he_cur = txt_file.read().strip()
                            print(hyperlink_path)
                            print(f"valuee isss: {he_cur}")
                        print(f"Fetched content from {hyperlink_path} and updated value.")
                        json_data[new_root_object] = try_parse_json(he_cur)
                    else:
                        json_data[new_root_object] = try_parse_json(he_cur)
                        print(f"Stored value in '{new_root_object}': {he_cur}")
                update_template.append(service_name)
                continue
 
            elif service_name in ['data', 'env']:
                pass  # Skip the key checks for 'data' and 'env'
 
            else:
                print(f"Error: Key is missing or empty in row {row_num}.")
                raise ValueError(f"Missing or empty key encountered in row {row_num}.")
 
 
        parsed_value = try_parse_json(he_cur)

        if parsed_value is None:
            print(f"Skipping row {row_num} due to None parsed_value")
            continue  # Skip processing this row if parsed_value is None

        if service_name in ['data', 'env']:
            handle_data_env(json_data, service_name, change_request, parsed_value)
            continue

        # General handling for other services with keys
        key_path = key.split('//')
        obj = json_data.setdefault(service_name, {})

        for k in key_path[:-1]:
            if not isinstance(obj, dict):
                obj = {}
            obj = obj.setdefault(k, {})

        final_key = key_path[-1]

        if key_path[0] == "env" or (len(key_path) > 1 and key_path[1] == "env"):
            if change_request == 'add':
                if final_key not in obj:
                    obj[final_key] = []
                obj[final_key].insert(-1, parsed_value)
                print(f"Added to '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'modify':
                for entry in obj[final_key]:
                    if entry['name'] == parsed_value['name']:
                        entry['value'] = parsed_value['value']
                        print(f"Modified '{final_key}' entry in '{service_name}': {parsed_value}")
                        break
            elif change_request == 'delete':
                if final_key in obj and parsed_value is not None:
                    print(final_key, obj)
                    print(f"Deleted entry from '{final_key}' in '{service_name}'.")
                    obj[final_key] = [
                        entry for entry in obj[final_key]
                        if not (isinstance(entry, dict) and 'name' in entry and parsed_value is not None and entry['name'] == parsed_value['name'])
                    ]

        else:
            if change_request == 'modify':
                obj[final_key] = parsed_value
                print(f"Modified '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'add':
                obj[final_key] = parsed_value
                print(f"Added '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'delete' and final_key in obj:
                del obj[final_key]
                print(f"Deleted '{final_key}' from '{service_name}'.")

        # Check for missing or empty value
        if he_cur is None or (isinstance(he_cur, str) and he_cur.strip() == ""):
            if change_request == "delete":
                continue
            else:
                raise ValueError(f"Missing or empty value encountered in row {row_num}.")

    return json_data

 
def extract_hyperlink_path(hyperlink_formula):
    """
    Extracts the file path from an Excel HYPERLINK formula.
    Assumes the formula is in the format: =HYPERLINK("path_to_file", "text")
    """
    # Extract the part inside the parentheses
    match = re.search(r'file:////(.*)', hyperlink_formula)
    if match:
        return match.group(1)
    return None
 
def save_json_to_file(json_data, output_file):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w') as f:
        json.dump(json_data, f, indent=4)
 
def process_json_data(data):
    if isinstance(data, list):
        return [process_json_data(item) for item in data]
    elif isinstance(data, dict):
        return {key: process_json_data(value) for key, value in data.items()}
    return data
 
def create_yaml_files_from_json(updated_output_file, output_folder):
    with open(updated_output_file, 'r') as json_file:
        json_data = json.load(json_file)
 
    os.makedirs(output_folder, exist_ok=True)
 
    for root_object, data in json_data.items():
        yaml_file_path = os.path.join(output_folder, f"{root_object}.yaml")
        processed_data = process_json_data(data)
 
        with open(yaml_file_path, 'w') as yaml_file:
            yaml.dump(processed_data, yaml_file, default_flow_style=False, sort_keys=False)
 
def move_line_containing_to_front(filename, line_to_reorder):
    """
    Moves a line containing the specified string to the beginning of the file.
 
    Args:
        filename (str): The path to the text file.
        line_to_reorder (str): The string to search for in the lines of the file.  The
                                  entire line containing this string will be moved
                                  to the front.
    """
    try:
        with open(filename, 'r+') as f:
            lines = f.readlines()
            found_line = None
            found_index = -1
 
            for i, line in enumerate(lines):
                if line_to_reorder in line:
                    found_line = line
                    found_index = i
                    break  # Stop searching after the first match
 
            if found_line is None:
                print(f"Error: No line containing '{line_to_reorder}' found in file.")
                return
 
            # Remove the found line
            del lines[found_index]
 
            # Insert the found line at the beginning
            lines.insert(0, found_line)
 
            # Write the modified content back to the file
            f.seek(0)
            f.writelines(lines)
            f.truncate()
 
        print(f"Successfully moved line containing '{line_to_reorder}' to the beginning of {filename}")
 
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")
 
def apply_sed_to_yaml(folder_path):
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".yaml"):
            file_path = os.path.join(folder_path, file_name)
            sed_command = f"sed -i -E -e 's/([[:space:]]+value:[[:space:]]*)(0[0-9]+)([[:space:]]*$)/\\1\"\\2\"\\3/' -e 's/([[:space:]]+value:[[:space:]]*)([A-Z])([[:space:]]*$)/\\1\"\\2\"\\3/' {file_path}"
 
            try:
                subprocess.run(sed_command, shell=True, check=True)
                print(f"Processed: {file_name}")
            except subprocess.CalledProcessError as e:
                print(f"Error processing {file_name}: {e}")
 
# def modify_deployment_yaml(folder, deleted_services):
#     """
#     Modifies the deployment.yaml file by removing the env blocks corresponding to deleted services.
#     """
#     temp_path = os.path.join(folder, "helm-charts","templates", "deployment.yaml")
 
#     """Removes the specified service blocks from the env section of the deployment.yaml."""
#     for service in deleted_services:
#         service = service.replace('-', '_')
#         pattern = rf"{{- with \.Values\.env\.{service} }}"
 
#         sed_command = [
#             "sed",
#             "-i.bak",
#             rf"/{pattern}/,+2d",
#             temp_path,
#         ]
#         result = subprocess.run(sed_command, capture_output=True, text=True, timeout=30)


def modify_deployment_yaml(folder, deleted_services):
    """
    Safely remove entire {{ with .Values.env.<service> }} ... {{ end }} blocks
    from deployment.yaml for deleted services.
    Counts block nesting to avoid leaving stray {{ end }}.
    """
    temp_path = os.path.join(folder, "helm-charts", "templates", "deployment.yaml")

    with open(temp_path, "r") as f:
        lines = f.readlines()

    new_lines = []
    skip_block = False
    block_depth = 0

    for line in lines:
        if not skip_block:
            # Check for start of a with block for a deleted service
            for service in deleted_services:
                service_name = service.replace("-", "_")
                if re.search(rf"{{[-]?\s*with\s+\.Values\.env\.{service_name}\s*}}", line):
                    skip_block = True
                    block_depth = 1
                    break

            if not skip_block:
                new_lines.append(line)
        else:
            # Inside a skipped block — track nesting
            if re.search(r"{{[-]?\s*(with|if|range)\b", line):
                block_depth += 1
            elif re.search(r"{{[-]?\s*end\s*}}", line):
                block_depth -= 1
                if block_depth == 0:
                    skip_block = False  # End of our block — don't append this end
            # Skip everything while inside the block

    with open(temp_path, "w") as f:
        f.writelines(new_lines)

 
def create_txt_file(excel_path, env, txt_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[env]  # Use sheet argument correctly
    max_row = ws.max_row
    values_to_be_added = set()
 
    for row in range(2, max_row + 1):  # Start from the second row
        name = ws.cell(row=row, column=1).value
        second_col_value = ws.cell(row=row, column=2).value
 
        if name is None:
            continue
 
        elif name in ('data', 'env'):
            continue
 
        elif second_col_value == "delete":
            # Skip writing this name to file
            continue
        else:
            values_to_be_added.add(name)
 
    with open(txt_path, 'w') as file:
        for i in values_to_be_added:
            file.write(f"{i}\n")
 
def update_txt_file_with_yaml_values(txt_file_path, image_tags_file ,yaml_folder_path):
    # Read the filenames from the .txt file
    with open(txt_file_path, 'r') as file:
        filenames = [line.strip() for line in file.readlines()]
 
    # For each filename in the txt file, find the corresponding YAML file
    for filename in filenames:
        yaml_file_path = os.path.join(yaml_folder_path, f"{filename}.yaml")
        if os.path.exists(yaml_file_path):
            with open(yaml_file_path, 'r') as yaml_file:
                yaml_content = yaml.safe_load(yaml_file)
                name_value = ''
                name_value = yaml_content.get('app', {}).get('name', None)
                if name_value:
                    # Replace the filename in the txt file with the name value from YAML
                    filenames[filenames.index(filename)] = f"{filename}:{name_value}"
 
    # Write the updated filenames back to the txt file
    with open(txt_file_path, 'w') as file:
        for filename in filenames:
            file.write(f"{filename}\n")
 
    image_tag_filenames = []
    order_for_deployment = ["secret-manager-v1", "config-cache-v1", "cs-secrets-service","user-status-service"]
 
 
    if os.path.exists(image_tags_file):
        with open(image_tags_file, 'r') as f:
            for line in f:
                parts = line.strip().split(':')  # Split each line by ':'
                if len(parts) >= 1:
                    filename = parts[0]  # The filename is the first part
                    image_tag_filenames.append(filename)
 
 
 
        with open(txt_file_path, 'r') as file:
            filenames = [line.strip() for line in file.readlines()]
            for filename in filenames:
                filename = filename.strip().split(":")[0]
                if filename in image_tag_filenames:
                    print("The file name is: ", filename)
                    # filenamee = filename.strip().split(":")[0]
                    if filename in order_for_deployment:
                        print("The filename is: ", filename)
                        move_line_containing_to_front(txt_file_path,filename)
    else:
        print("Image tags not available")
 
def update_meta_sheet(lower_env, higher_env, promote_branch, repo_url):
    repo_path = tempfile.mkdtemp()

    try:
        repo_url = inject_git_token(repo_url)

        clone_repo_and_checkout(repo_url, 'master', repo_path)

        if not is_base_branch_exists(repo_url, promote_branch):
            raise ValueError(f"Branch {promote_branch} does not exist in repository")

        print("Step1")

        excel_path = os.path.join(repo_path, 'meta-sheet.xlsx')
        wb = load_workbook(excel_path)
        ws = wb.active
        if os.path.exists(excel_path):
            print("path exists: ", excel_path)
        headers = [cell.value for cell in ws[1]]
        lower_col = headers.index(lower_env) + 1
        higher_col = headers.index(higher_env) + 1

        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[lower_col - 1].value == promote_branch:
                if row[higher_col - 1].value == promote_branch:
                    print("already updated")
                else:
                    row[higher_col - 1].value = promote_branch
                updated = True
                break
        print("Step2")
        if not updated:
            raise ValueError(f"{promote_branch} not found in {lower_env} column")

        wb.save(excel_path)
        print("saved excel file")

        print("Checking if there are changes to commit...")
        status = subprocess.run(['git', '-C', repo_path, 'status', '--porcelain'], capture_output=True, text=True)
        if not status.stdout.strip():
            print("No changes to commit.")
            return True

        configure_git_user(repo_path)
        stage_specific_files_commit_and_push(
            repo_path,
            'master',
            f'Promoted {promote_branch} from {lower_env} to {higher_env}',
            ['meta-sheet.xlsx'],
            pull_before_push=False,
        )

    except subprocess.CalledProcessError as e:
        print(f"Git operation failed: {e}")
        print(f"Return code: {e.returncode}")
        print(f"Command: {e.cmd}")
        print(f"Output: {e.output}")
        print(f"Stderr: {e.stderr}")
        return False
 
def fetch_branches(repo_url, lower_env, higher_env):

    def find_column_index(sheet, env_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == env_name:
                return col
        raise ValueError(f"Environment '{env_name}' not found in header")

    def find_last_updated_branch(sheet, col_idx):
        for row in range(sheet.max_row, 1, -1):
            val = sheet.cell(row=row, column=col_idx).value
            if val is not None and val != 'X':
                return val, row
        raise ValueError(f"No branch found in column index {col_idx}")

    repo_url = inject_git_token(repo_url)

    with tempfile.TemporaryDirectory() as tmpdirname:
        clone_single_branch_and_checkout(repo_url, 'master', tmpdirname, depth=1)

        file_path = os.path.join(tmpdirname, 'meta-sheet.xlsx')
        wb = load_workbook(file_path)
        sheet = wb.active

        lower_col = find_column_index(sheet, lower_env)
        higher_col = find_column_index(sheet, higher_env)

        lower_branch, _ = find_last_updated_branch(sheet, lower_col)
        higher_branch, _ = find_last_updated_branch(sheet, higher_col)

    return lower_branch, higher_branch
 
 
def main():
 
    repos_info = {
        'promotion-repo': sys.argv[1]
    }
 
    lower_env = sys.argv[4]
    higher_env = sys.argv[5]
 
    # promote_branch_x, promote_branch_x_1 = fetch_branches(sys.argv[1],lower_env, higher_env)
 
 
    for repo_name, repo_url in repos_info.items():
        promote_branch_x_1 = sys.argv[2]
        promote_branch_x = sys.argv[3]
 
        target_folder_x_1 = os.path.join(os.getcwd(), "generate-config", "promotion-x-1", f"{repo_name}")
        target_folder_x = os.path.join(os.getcwd(), "generate-config", "promotion-x", f"{repo_name}")
 
    repo_url = inject_git_token(repo_url)
    clone_single_branch_and_checkout(repo_url, promote_branch_x_1, target_folder_x_1)
    clone_single_branch_and_checkout(repo_url, promote_branch_x, target_folder_x)
 
    lower_env = sys.argv[4]
    higher_env = sys.argv[5]
 
    release_note_path = os.path.join(target_folder_x, "helm-charts", f"{higher_env}-values", "app-values", f"release_note")
    print("Path is :: ", os.path.join(target_folder_x, "helm-charts", f"{higher_env}-values", f"release_note"))
 
    print("These are the values:", repo_url, promote_branch_x, lower_env, target_folder_x)
    foldernames = os.listdir(target_folder_x)
 
    for foldername in foldernames:
        print("These are the folder: ",foldername)
        if foldername == "helm-charts":
            for filename in os.listdir(release_note_path):
                if filename.startswith("release-note") and "verified" in filename:
                    excel_file_path =  os.path.join(release_note_path,filename)
                    print("Excel file path: ", excel_file_path)
                else:
                    print("Release note does not exist")
 
 
            print("hiighherrr: ",higher_env)
            sheet = get_sheets_with_values(excel_file_path, higher_env)
            print(f"Promoting the values in env: {sheet} of {foldername}")
            folder_path = os.path.join(target_folder_x_1, "helm-charts", f"{sheet}-values", "app-values" )
            initial_output_file = os.path.join(target_folder_x_1, "helm-charts", f"{sheet}-values", "app-values", f"config-{sheet}.json")
            updated_output_file = os.path.join(target_folder_x, "helm-charts", f"{sheet}-values", "app-values", f"config-{sheet}.json")
            output_folder = os.path.join(target_folder_x, "helm-charts", f"{sheet}-values", "app-values" )
            txt_file_path = os.path.join(target_folder_x, "helm-charts", f"{sheet}-values", "app-values", f"{sheet}.txt" )
            image_tag_file_path = os.path.join(target_folder_x, f"upgrade-services.txt")
 
            json_data = read_yaml_files_to_json(folder_path)
            save_json_to_file(json_data, initial_output_file)
            updated_json = apply_changes_to_json(json_data, excel_file_path, sheet, lower_env, higher_env)
            insert_hardcoded_value(target_folder_x,update_template)
            modify_deployment_yaml(target_folder_x,deleted_services)
            save_json_to_file(updated_json, updated_output_file)
            create_yaml_files_from_json(updated_output_file, output_folder)
            apply_sed_to_yaml(output_folder)
            create_txt_file(excel_file_path,sheet,txt_file_path)
            # update_txt_file_with_yaml_values(txt_file_path,image_tag_file_path,output_folder)
 
    configure_git_user(target_folder_x)
    stage_commit_and_push(
        target_folder_x,
        sys.argv[3],
        f'Config files generated: {sys.argv[3]}',
    )

    update_meta_sheet(lower_env, sheet, promote_branch_x, sys.argv[1])
 
 
    target_folder_x = os.path.dirname(target_folder_x)
    target_folder_x_1 = os.path.dirname(target_folder_x_1)
 
if __name__ == "__main__":
    main()
 
